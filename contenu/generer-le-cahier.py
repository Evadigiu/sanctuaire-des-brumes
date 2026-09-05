# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ARIAL   = "Arial"
JAUNE   = PatternFill("solid", fgColor="FFF2CC")   # a remplir
GRIS    = PatternFill("solid", fgColor="EDEDED")   # ne pas toucher
ENTETE  = PatternFill("solid", fgColor="2C3247")
ROUGE   = PatternFill("solid", fgColor="FCE4E4")   # question a trancher
VERT    = PatternFill("solid", fgColor="E6F0E6")   # exemple
BORD    = Border(*[Side(style="thin", color="BBBBBB")]*4)

def h(ws, row, cols, widths):
    for i, (c, w) in enumerate(zip(cols, widths), start=1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.font = Font(name=ARIAL, bold=True, size=10, color="FFFFFF")
        cell.fill = ENTETE
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORD
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 30

# ============================================================
# LE PARCOURS, tel que je le lis sur le schema
# id, nom, type, lieu suppose, lettre, ordre A, ordre B
# ============================================================
ETAPES = [
 ("E01","Jerry, les consignes","Temoin","Jardin des pivoines","", 1, 1),
 ("E02","La collegue soigneuse","Temoin","Bureau des soignants","", 2, 6),
 ("E03","Indice : les cameras de surveillance","Indice","Salle de seminaire","", 3, 7),
 ("E04","Epreuve : le plan des deplacements","Epreuve","", "", 4, 8),
 ("E05","Indice : les jumelles, la carcasse","Indice","", "", 5, 9),
 ("E06","Epreuve : le chapeau de l'epouvantail","Epreuve","Enclos des loups","R", 6, 10),
 ("E07","La passante","Temoin","", "", 7, 11),
 ("E08","Quizz sonore, Sabri & Arez","Epreuve","Enclos des lynx","I", 8, 12),
 ("E09","Le veterinaire","Temoin","", "", 9, 13),
 ("E10","Indice : le panneau d'empreinte","Indice","", "S", 10, 14),
 ("E11","Greg, version sens A","Temoin","", "", 11, None),
 ("E11B","Greg B, version sens B","Temoin","", "", None, 2),
 ("E12","Bill","Temoin","", "", 12, 3),
 ("E13","Indice : le bureau de Greg","Indice","Bureau de Greg","I", 13, 5),
 ("E14","Epreuve botanique, la tulipe","Epreuve","Jardin des tulipes","", 14, 4),
 ("E15","Jardin des Iris, le code de la serre","Enigme","Jardin des Iris","", 15, 15),
 ("E16","Le botaniste","Temoin","La serre","", 16, 16),
 ("E17","La resolution finale","Final","", "", 17, 17),
]

def suivant(ordre_key):
    """Calcule l'etape suivante dans un sens donne."""
    idx = 5 if ordre_key == "A" else 6
    ordonne = sorted([e for e in ETAPES if e[idx] is not None], key=lambda e: e[idx])
    res = {}
    for i, e in enumerate(ordonne):
        res[e[0]] = ordonne[i+1][0] + " - " + ordonne[i+1][1] if i+1 < len(ordonne) else "FIN DU PARCOURS"
    return res

SUIV_A, SUIV_B = suivant("A"), suivant("B")

wb = Workbook()

# ============================================================
# ONGLET 1 : LISEZ-MOI
# ============================================================
ws = wb.active
ws.title = "Lisez-moi"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 105

def p(row, texte, gras=False, taille=10, couleur="000000", fill=None):
    c = ws.cell(row=row, column=2, value=texte)
    c.font = Font(name=ARIAL, bold=gras, size=taille, color=couleur)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    if fill: c.fill = fill
    ws.row_dimensions[row].height = None if len(texte) < 90 else 15*(len(texte)//90 + 1)
    return c

p(2,  "LE SANCTUAIRE DES BRUMES", True, 16, "2C3247")
p(3,  "Cahier de contenu du jeu", False, 12, "8B7355")
p(5,  "A QUOI SERT CE FICHIER", True, 11, "2C3247")
p(6,  "C'est ici que s'ecrit tout le texte du jeu. Une fois rempli, les pages du site sont fabriquees "
      "automatiquement a partir de ce fichier. Tu n'ecris jamais dans le code, tu ecris seulement ici.")
p(7,  "Consequence importante : ce fichier devient la reference. Si un texte doit changer, il change ici, "
      "puis les pages sont refabriquees. Ne corrige jamais un texte directement dans une page du site, "
      "il serait ecrase a la fabrication suivante.")
p(9,  "LES TROIS ONGLETS, EN BAS DE L'ECRAN", True, 11, "2C3247")
p(10, "Lisez-moi     ....  cette page, plus les questions en attente de ta reponse")
p(11, "Parcours      ....  la structure du jeu : les etapes, leur ordre dans chaque sens, les lieux")
p(12, "Textes        ....  les textes a ecrire, un par ligne")
p(14, "CE QUE TU DOIS REMPLIR", True, 11, "2C3247")
c = p(15, "     Les cases sur fond creme sont a remplir par toi.")
ws.cell(row=15, column=2).fill = JAUNE
p(16, "     Les cases sur fond gris sont calculees ou fixees, tu peux les ignorer.")
ws.cell(row=16, column=2).fill = GRIS
p(17, "     Les cases sur fond rose sont des questions qui attendent ta reponse.")
ws.cell(row=17, column=2).fill = ROUGE
p(19, "AVANCEMENT", True, 11, "2C3247")
p(20, "86 textes a ecrire au total, repartis sur les 17 etapes.", True, 11, "2C3247")
p(21, "Astuce pour suivre ton avancement : clique sur la lettre F en haut de la colonne des textes, "
      "dans l'onglet Textes. En bas a droite de l'ecran, Excel affiche le nombre de cases remplies.", False, 10, "8B7355")

p(23, "QUATRE QUESTIONS AVANT QUE JE FABRIQUE QUOI QUE CE SOIT", True, 11, "2C3247")
questions = [
 "1. COMBIEN DE QR CODES ? Ton schema montre 17 etapes. La base de donnees n'en connait que 12. "
 "Certaines etapes sont-elles des ecrans successifs d'une meme borne (on scanne une fois, puis on "
 "enchaine indice puis epreuve), ou chaque etape a-t-elle son propre QR code dans le parc ? "
 "Reponds dans l'onglet Parcours, colonne \"Borne QR ?\".",

 "2. LES LETTRES DE L'ENIGME. En sens B, le joueur ramasse I, R, I, S dans cet ordre : le mot est "
 "deja ecrit, l'enigme du Jardin des Iris se resout sans reflechir. En sens A il ramasse R, I, S, I "
 "et doit chercher. Veux-tu permuter deux lettres pour reequilibrer ?",

 "3. LE CODE DE LA SERRE. Le joueur tape-t-il IRIS sur le site, ou sur un cadenas physique ? "
 "Si c'est sur le site, c'est une mecanique a construire, dis-le moi.",

 "4. LES DEUX PISTES DE LA COLLEGUE SOIGNEUSE. Sur ton schema, elle ouvre deux branches : les cameras "
 "puis le plan d'un cote, les jumelles puis l'enclos des loups de l'autre. Les deux sont-elles "
 "obligatoires, ou la premiere est-elle facultative ? J'ai suppose qu'on les fait l'une apres l'autre.",
]
r = 24
for q in questions:
    cc = p(r, q, False, 10)
    cc.fill = ROUGE
    cc.border = BORD
    ws.row_dimensions[r].height = 15 * (len(q)//95 + 2)
    r += 2

# ============================================================
# ONGLET 2 : PARCOURS
# ============================================================
ws = wb.create_sheet("Parcours")
ws.sheet_view.showGridLines = False
ws["A1"] = "LA STRUCTURE DU JEU. Les colonnes creme sont a completer par toi."
ws["A1"].font = Font(name=ARIAL, bold=True, size=11, color="2C3247")
h(ws, 2,
  ["Code","Nom de l'etape","Type","Lieu dans le parc","Borne QR ?","Lettre",
   "Ordre sens A","Suite en sens A","Ordre sens B","Suite en sens B","Remarques"],
  [8, 34, 10, 24, 12, 8, 12, 34, 12, 34, 40])

r = 3
for eid, nom, typ, lieu, lettre, oa, ob in ETAPES:
    vals = [eid, nom, typ, lieu, "", lettre,
            oa if oa else "—", SUIV_A.get(eid, "—"),
            ob if ob else "—", SUIV_B.get(eid, "—"), ""]
    for i, v in enumerate(vals, start=1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = Font(name=ARIAL, size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.border = BORD
        if i in (4, 5, 11): c.fill = JAUNE      # a remplir
        else: c.fill = GRIS
    if eid in ("E11", "E11B"):
        ws.cell(row=r, column=11, value="Le temoin dedouble : un texte different selon le sens.").font = Font(name=ARIAL, size=9, italic=True)
        ws.cell(row=r, column=11).fill = JAUNE
    if eid == "E14":
        ws.cell(row=r, column=11, value="NOUVELLE etape, absente de la base de donnees. A ajouter.").font = Font(name=ARIAL, size=9, italic=True)
        ws.cell(row=r, column=11).fill = JAUNE
    if eid in ("E04","E05","E10"):
        ws.cell(row=r, column=11, value="Etape lue sur le schema, absente de la base de donnees. A ajouter si c'est une borne.").font = Font(name=ARIAL, size=9, italic=True)
        ws.cell(row=r, column=11).fill = JAUNE
    ws.row_dimensions[r].height = 30
    r += 1
ws.freeze_panes = "C3"

# ============================================================
# ONGLET 3 : TEXTES
# ============================================================
ws = wb.create_sheet("Textes")
ws.sheet_view.showGridLines = False
ws["A1"] = "LES TEXTES DU JEU. Ecris uniquement dans la colonne F, sur fond creme. La ligne verte est un exemple, supprime-la quand tu auras compris."
ws["A1"].font = Font(name=ARIAL, bold=True, size=11, color="2C3247")
h(ws, 2, ["Code","Etape","Sens","Ecran","Ce qu'il faut ecrire ici","TEXTE (a ecrire)","Remarques"],
   [10, 30, 10, 26, 44, 60, 30])

ECRANS = [
 ("1. Accueil apres le scan", "les deux",
  "La phrase d'accroche affichee juste apres le scan, au-dessus du bouton. Une ou deux lignes, pas plus : le groupe est debout, il n'a pas envie de lire."),
 ("2. Legende du media",      "les deux",
  "La phrase sous la video ou l'audio. Elle dit qui parle et de quoi. Une ligne."),
 ("3. Texte ou epreuve",      "les deux",
  "Le coeur de l'etape : le temoignage ecrit, la consigne de l'epreuve, ou l'indice. Aussi long que necessaire."),
 ("4. Ou aller ensuite",      "sens A",
  "La phrase qui envoie le joueur a l'etape suivante DU SENS A. C'est elle qui fait vivre les deux sens."),
 ("4. Ou aller ensuite",      "sens B",
  "La phrase qui envoie le joueur a l'etape suivante DU SENS B."),
]

r = 3
# ligne d'exemple
exemple = ["E02","La collegue soigneuse","les deux","1. Accueil apres le scan",
           "La phrase d'accroche affichee juste apres le scan...",
           "Elle vous attend depuis ce matin. Elle n'a pas dormi.",
           "EXEMPLE, a supprimer"]
for i, v in enumerate(exemple, start=1):
    c = ws.cell(row=r, column=i, value=v)
    c.font = Font(name=ARIAL, size=10, italic=True)
    c.fill = VERT
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.border = BORD
ws.row_dimensions[r].height = 30
r += 1

for eid, nom, typ, lieu, lettre, oa, ob in ETAPES:
    for libelle, sens, consigne in ECRANS:
        # Greg dedouble : une seule orientation chacun
        if eid == "E11"  and sens == "sens B": continue
        if eid == "E11B" and sens == "sens A": continue
        if eid == "E17"  and libelle.startswith("4."): continue
        vals = [eid, nom, sens, libelle, consigne, "", ""]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = Font(name=ARIAL, size=10)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = BORD
            if i == 6: c.fill = JAUNE
            elif i == 7: c.fill = JAUNE
            else: c.fill = GRIS
        ws.row_dimensions[r].height = 30
        r += 1
ws.freeze_panes = "C3"
print("lignes de texte a remplir :", r - 4)

import os
out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sanctuaire-cahier-de-contenu.xlsx")
wb.save(out)
print("ecrit :", out)
